import os
from datetime import datetime, timedelta
from tkinter import Tk, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Side, Alignment


# Función para obtener el nombre de la hoja según el criterio proporcionado
def obtener_nombre_hoja():
    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Retroceder al domingo anterior
    domingo_anterior = fecha_actual - timedelta(days=fecha_actual.weekday() + 1)

    # Avanzar al sábado siguiente
    sabado_siguiente = domingo_anterior + timedelta(days=6)

    # Mapear nombres de meses en inglés a español
    meses_en_espanol = {
        'Jan': 'Ene',
        'Feb': 'Feb',
        'Mar': 'Mar',
        'Apr': 'Abr',
        'May': 'May',
        'Jun': 'Jun',
        'Jul': 'Jul',
        'Aug': 'Ago',
        'Sep': 'Sep',
        'Oct': 'Oct',
        'Nov': 'Nov',
        'Dec': 'Dic'
    }

    # Verificar si el domingo anterior y el sábado siguiente están en diferentes meses
    if domingo_anterior.month != sabado_siguiente.month:
        nombre_hoja = f"{domingo_anterior.day} {meses_en_espanol[datetime(domingo_anterior.year, domingo_anterior.month, 1).strftime('%b')]} - {sabado_siguiente.day} {meses_en_espanol[datetime(sabado_siguiente.year, sabado_siguiente.month, 1).strftime('%b')]}"
    else:
        nombre_hoja = f"{domingo_anterior.day} - {sabado_siguiente.day} {meses_en_espanol[datetime(sabado_siguiente.year, sabado_siguiente.month, 1).strftime('%b')]}"

    return nombre_hoja

# Función para crear una hoja en el archivo Excel seleccionado
def crear_hoja_excel():
    # Crear una ventana de selección de archivo
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="Selecciona un archivo Excel", filetypes=[("Archivos Excel", "*.xlsx")])

    # Verificar si se seleccionó un archivo
    if file_path:
        # Cargar el archivo Excel existente o crear uno nuevo
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        # Obtener el nombre de la hoja según el criterio
        nombre_hoja = obtener_nombre_hoja()

        # Crear una nueva hoja con el nombre obtenido
        sheet = workbook.create_sheet(title=nombre_hoja)

        # Agregar contenido y formato a las celdas
         # Combinar celdas A1:F1
        sheet.merge_cells('A1:F1')
                
        # Establecer el ancho de las columnas A a F a 100 píxeles
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            sheet.column_dimensions[col_letter].width = 15
        # sheet.column_dimensions['A'].height = 30
        # Celdas A1:F1
        cell_A1 = sheet['A1']
        cell_A1.value = "5 ESQUINAS - MARIA QUISPE"
        cell_A1.font = Font(color="000000", bold=True)
        cell_A1.font = Font(size=24)
        # cell_A1.alignment = Alignment(horizontal='center', vertical='center')
        cell_A1.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


        # Celdas A2:F2
        encabezados = ["N*GUIA", "FECHA", "CANT.", "PRODUCTOS", "P. UNIT.", "MONTO TOTAL"]
        for col_num, encabezado in enumerate(encabezados, 1):
            cell = sheet[get_column_letter(col_num) + '2']
            cell.value = encabezado
            cell.font = Font(color="000000", bold=True)
            cell.fill = PatternFill(start_color="ef6b26", end_color="ef6b26", fill_type="solid")

        # Llenar 20 filas de datos
        for row_num in range(3, 23):
            sheet[f'F{row_num}'] = f'=$E{row_num}*C{row_num}'  # Fórmula para calcular Monto Total

        # Calcular el Total en la última fila de la columna F
        sheet[f'E23'] = f'TOTAL'
        sheet[f'F23'] = f'=SUM(F3:F22)'
        sheet[f'F23'].font = Font(color="ff0000", bold=True)
        # Establecer el formato de número como moneda peruana para la celda F23
        sheet['F23'].number_format = '"S/ "#,##0.00'

        for row in sheet['E23:F23']:
                    for cell in row:
                        cell.border = cell.border.copy(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )

        for row in sheet['A1:F22']:
            for cell in row:
                cell.border = cell.border.copy(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        for row in sheet.iter_rows(min_row=1, max_row=23, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el archivo Excel actualizado
        workbook.save(file_path)
        print(f"Se ha creado la hoja '{nombre_hoja}' en el archivo '{file_path}'.")

# Llamar a la función para crear la hoja en el archivo Excel seleccionado
crear_hoja_excel()